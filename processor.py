"""
processor.py
============
출고현황 / 반품현황 → 오프라인 판매 업로드 엑셀 생성 핵심 로직
XLS 파일은 xlrd 로 직접 처리 (LibreOffice 불필요)
"""

import re
import io
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 컬럼 인덱스 (0-based) ──────────────────────────────────────
COL_DATE     = 11
COL_STORE_CD = 13
COL_STORE    = 14
COL_STYLE    = 15
COL_COLOR    = 17
COL_SIZE     = 18
COL_QTY      = 19
COL_PRICE    = 20
COL_AMOUNT   = 21

# ── 스타일 상수 ───────────────────────────────────────────────
COLOR_HEADER = "FFD9E1F2"
COLOR_TOTAL  = "FFFCE4D6"
COLOR_BORDER = "FF8EA9C1"


# ═══════════════════════════════════════════════════════════════
# 파일 읽기
# ═══════════════════════════════════════════════════════════════

def read_file(file_obj) -> pd.DataFrame:
    """
    xls / xlsx 파일 객체(혹은 경로)를 DataFrame 으로 읽기.
    file_obj: 파일 경로(str) 또는 BytesIO / 업로드 파일 객체
    """
    # 파일명 또는 객체에서 확장자 판별
    if isinstance(file_obj, str):
        ext = file_obj.rsplit(".", 1)[-1].lower()
        data = file_obj
    else:
        name = getattr(file_obj, "name", "")
        ext  = name.rsplit(".", 1)[-1].lower()
        data = io.BytesIO(file_obj.read())

    if ext == "xls":
        try:
            import xlrd  # Streamlit Cloud / pip 환경에서 설치됨
            df = pd.read_excel(data, engine="xlrd", header=None)
        except ImportError:
            raise RuntimeError(
                "xlrd 패키지가 필요합니다.\n"
                "터미널에서 실행: pip install xlrd"
            )
    elif ext == "xlsx":
        df = pd.read_excel(data, engine="openpyxl", header=None)
    else:
        raise ValueError(f"지원하지 않는 파일 형식: .{ext}  (xls 또는 xlsx 만 가능)")

    return df


# ═══════════════════════════════════════════════════════════════
# 데이터 처리
# ═══════════════════════════════════════════════════════════════

def _parse_money(val) -> int:
    """'69,900  /  '-39,900  →  정수로 변환"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0
    s = str(val).strip().replace("'", "").replace(",", "").replace(" ", "")
    try:
        return int(float(s))
    except ValueError:
        return 0


def _parse_color(color_str: str) -> str:
    """'03:네이비'  →  '03'  (숫자 코드만 추출)"""
    m = re.match(r"^(\d{2}):", str(color_str).strip())
    return m.group(1) if m else str(color_str).strip()


def _make_key(row):
    style    = str(row.iloc[COL_STYLE]).strip()
    color    = _parse_color(str(row.iloc[COL_COLOR]))
    size_raw = row.iloc[COL_SIZE]
    size     = str(int(size_raw)) if pd.notna(size_raw) else ""
    store_cd = str(row.iloc[COL_STORE_CD]).strip()
    return (store_cd, style, color, size)


def aggregate(df: pd.DataFrame) -> dict:
    """
    (매장코드, 품번, 컬러코드, 사이즈) 키로 수량·금액 집계.
    헤더 3행 + 마지막 합계행 제외.
    """
    result = {}
    data   = df.iloc[3:-1]

    for _, row in data.iterrows():
        # 날짜 없는 행(소계 등) 스킵
        if pd.isna(row.iloc[COL_DATE]):
            continue

        key      = _make_key(row)
        qty      = _parse_money(row.iloc[COL_QTY])
        amount   = _parse_money(row.iloc[COL_AMOUNT])
        price    = _parse_money(row.iloc[COL_PRICE])   # 출고단가
        store_nm = str(row.iloc[COL_STORE]).strip()

        if key not in result:
            result[key] = {"매장명": store_nm, "수량": 0, "금액": 0, "단가": price}
        result[key]["수량"] += qty
        result[key]["금액"] += amount
        # 단가는 0이 아닌 첫 값으로 확정
        if result[key]["단가"] == 0 and price != 0:
            result[key]["단가"] = price

    return result


def get_last_return_date(df: pd.DataFrame) -> datetime.datetime:
    """반품 파일의 마지막(최대) 날짜 반환"""
    dates = [
        d for d in df.iloc[3:-1].iloc[:, COL_DATE]
        if hasattr(d, "date")
    ]
    if not dates:
        raise ValueError("반품 파일에서 날짜 데이터를 읽을 수 없습니다.")
    return max(dates)


def compute_net(out_agg: dict, ret_agg: dict) -> list:
    """출고 - 반품 = 순 판매 목록"""
    rows = []
    for key, out in out_agg.items():
        ret     = ret_agg.get(key, {"수량": 0, "금액": 0})
        net_qty = out["수량"] - abs(ret["수량"])
        net_amt = out["금액"] - abs(ret["금액"])

        store_cd, style, color, size = key
        barcode = f"{style}{color}{size}"

        rows.append({
            "store_cd": store_cd,
            "store_nm": out["매장명"],
            "barcode":  barcode,
            "style":    style,
            "color":    color,
            "size":     size,
            "qty":      net_qty,
            "amount":   net_amt,
            "price":    out.get("단가", 0),
        })
    return rows


# ═══════════════════════════════════════════════════════════════
# 엑셀 출력
# ═══════════════════════════════════════════════════════════════

def _border_style():
    s = Side(border_style="thin", color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)


def _set_cell(cell, value=None, bold=False, fill_color=None,
              align="center", number_format=None):
    if value is not None:
        cell.value = value
    cell.font      = Font(bold=bold, name="맑은 고딕", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _border_style()
    if fill_color:
        cell.fill = _fill(fill_color)
    if number_format:
        cell.number_format = number_format


HEADERS = [
    ("A", "날짜"),         ("B", "매장코드"),    ("C", "매장명"),
    ("D", "상품유형"),      ("E", "공급처코드"),   ("F", "공급처명"),
    ("G", "상품차수"),      ("H", "상품연도"),     ("I", "대표시즌"),
    ("J", "시즌"),          ("K", "브랜드"),       ("L", "명"),
    ("M", "품목"),          ("N", "명"),           ("O", "품목군"),
    ("P", "상품코드(바코드)"), ("Q", "스타일"),    ("R", "컬러"),
    ("S", "사이즈"),        ("T", "최초소비자가"), ("U", "현 판매가"),
    ("V", "매입원가"),      ("W", "판매수량"),     ("X", "소가액"),
    ("Y", "매출액"),        ("Z", "실 판매액"),    ("AA", "원가액"),
    ("AB", "비고"),
]


def build_excel(rows: list, last_date: datetime.datetime) -> bytes:
    """
    결과 데이터를 오프라인 판매 업로드 엑셀 양식으로 만들어 bytes 반환.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "판매업로드"

    # ── 1행: 구분 헤더 ────────────────────────────────────────
    ws.merge_cells("A1:C1")
    ws.merge_cells("D1:O1")
    ws.merge_cells("T1:V1")
    for coord, label in [("A1", "구분"), ("D1", "상품속성"), ("T1", "가격")]:
        _set_cell(ws[coord], value=label, bold=True, fill_color=COLOR_HEADER)

    # ── 2행: 컬럼명 ──────────────────────────────────────────
    for col_letter, name in HEADERS:
        _set_cell(ws[f"{col_letter}2"], value=name, bold=True, fill_color=COLOR_HEADER)

    # ── 3행: 합계 ────────────────────────────────────────────
    data_last = 3 + len(rows)   # 마지막 데이터 행 번호
    ws["A3"] = "합계"
    ws["W3"] = f"=SUM(W4:W{data_last})"
    ws["Z3"] = f"=SUM(Z4:Z{data_last})"
    for col_letter, _ in HEADERS:
        cell = ws[f"{col_letter}3"]
        _set_cell(cell, bold=True, fill_color=COLOR_TOTAL)
    ws["W3"].number_format = "#,##0"
    ws["Z3"].number_format = "#,##0"

    # ── 4행~: 데이터 ─────────────────────────────────────────
    date_str = last_date.strftime("%Y-%m-%d")

    for i, r in enumerate(rows):
        row_num = 4 + i
        vals = {
            "A": date_str,
            "B": r["store_cd"],
            "C": r["store_nm"],
            "P": r["barcode"],
            "Q": r["style"],
            "R": r["color"],
            "S": r["size"],
            "U": r["price"],
            "W": r["qty"],
            "Z": r["amount"],
        }
        for col_letter, _ in HEADERS:
            cell = ws[f"{col_letter}{row_num}"]
            value = vals.get(col_letter)
            _set_cell(cell, value=value,
                      align="center" if col_letter in ("A","B","R","S","U","W","Z") else "left")

        ws[f"U{row_num}"].number_format = "#,##0"
        ws[f"W{row_num}"].number_format = "#,##0"
        ws[f"Z{row_num}"].number_format = "#,##0"

    # ── 열 너비 ──────────────────────────────────────────────
    for col_letter, width in [
        ("A", 14), ("B", 12), ("C", 18),
        ("P", 24), ("Q", 16), ("R", 8), ("S", 8),
        ("W", 10), ("Z", 16),
    ]:
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20

    # bytes 로 반환
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════
# 메인 실행 함수 (Streamlit / CLI 공용)
# ═══════════════════════════════════════════════════════════════

def run(out_file, ret_file) -> tuple[bytes, dict]:
    """
    Parameters
    ----------
    out_file : 파일 경로(str) 또는 Streamlit UploadedFile
    ret_file : 파일 경로(str) 또는 Streamlit UploadedFile

    Returns
    -------
    (엑셀 bytes, 요약 dict)
    """
    df_out = read_file(out_file)
    df_ret = read_file(ret_file)

    out_agg   = aggregate(df_out)
    ret_agg   = aggregate(df_ret)
    last_date = get_last_return_date(df_ret)
    rows      = compute_net(out_agg, ret_agg)

    excel_bytes = build_excel(rows, last_date)

    summary = {
        "날짜":      last_date.strftime("%Y-%m-%d"),
        "항목수":    len(rows),
        "총판매수량": sum(r["qty"]    for r in rows),
        "총실판매액": sum(r["amount"] for r in rows),
    }
    return excel_bytes, summary
